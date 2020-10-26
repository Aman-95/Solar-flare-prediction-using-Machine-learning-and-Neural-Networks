import openpyxl as op
## For fixing dates
wb=op.load_workbook("dataset_2020_new.xlsx")
s1=wb.get_sheet_by_name("Sheet 1")


for i in range(2,135):
    c1 = s1.cell(row=i, column=3)
    c2 = s1.cell(row=i, column=2)
    c3 = s1.cell(row=i, column=4)
    y=c3.value
    list1 = y.split(" ")
    print(list1," ",i)
    w1 = list1[0]
    w2= list1[1]
    
    c1.value=w1
    c2.value=w2
wb.save("dataset_2020_new.xlsx")
'''
for i in range(2,135):
    months=["January","February","March","April","May","June","July","August","September","October","November","December"]
    c=s1.cell(row=i,column=4)
    c1=c.value
    for count, m in enumerate(months,1):
        if c.value==m:
            c.value=count
    #print(c1," ",c.value)
wb.save("dataset_2020_new.xlsx")
'''