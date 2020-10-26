import openpyxl as op
import numpy as np
wb=op.load_workbook("Final_dataset - nearest vector filling - to compare - new.xlsx")
s1=wb.get_sheet_by_name("Sheet1")

#=======================================================================================================================
# first replace every None value by star
for i in range(3,8382):
    for j in range(4,50):
        c=s1.cell(row=i,column=j)
        if c.value==None or c.value==" " or c.value=="- " or c.value=="* " or c.value=="nk"or c.value=="-":
            print(f"({i},{j})")
            c.value="*"
wb.save("Final_dataset - nearest vector filling_trial2_modified.xlsx")
#==============================================================================w=========================================

"""
def distance(vec1,vec2,col):
    dist = 0
    for j3 in range(4,50):
        if j3 in col:
            continue
        else:
            dist=dist+((vec1[j3-5]-vec2[j3-5])/(vec1[j3-5]+1))**2
    return dist**(0.5)

def nearest_vector(i1):
    v1=np.linspace(4,49,46)
    col_to_skip=[]
    for k1 in range(4,50):
        c=s1.cell(row=i1,column=k1)
        if c.value=="*":
            col_to_skip.append(k1-5)
            v1[k1-5]=0
        else:
            v1[k1-5]=c.value
    min=999999999999999999
    min_vec=-1

    if i1<183:
        start=i1
        stop=i1+361
    elif i1>=183 and i1<=8382-180:
        start=i1-180
        stop=i1+181
    else:
        start=i1-360
        stop=i1+1

    for k2 in range(start,stop):
        if k2==i1:
            continue
        v2=np.linspace(4,49,46)
        skip=0
        for j2 in range(4,50):
            c=s1.cell(row=k2,column=j2)
            if c.value=="*":
                skip=1
                break
            v2[j2 - 5] = c.value
        if skip==1:
            continue
        else:
            score=distance(v1,v2,col_to_skip)
            if score < min :
                min=score
                min_vec=k2

    for k3 in col_to_skip:
        n=k3+5
        c1=s1.cell(row=i1,column=n)
        c2=s1.cell(row=min_vec,column=n)
        c1.value=c2.value
    return min_vec

#near_vect_index=nearest_vector(i)
#print(i,near_vect_index)
#(5410,11)
#(7391,11)

for i in range(3,8382):
    for j in range(4,50):
        c=s1.cell(row=i,column=j)
        if c.value=="*":
            near_vect_index=nearest_vector(i)
            print(i,near_vect_index)
            break
wb.save("Final_dataset - nearest vector filling_trial1_modified.xlsx")
"""