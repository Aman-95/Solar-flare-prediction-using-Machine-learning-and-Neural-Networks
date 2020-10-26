import openpyxl as op
## For fixing k indices
wb=op.load_workbook("dataset_2020_new.xlsx")
s1=wb.get_sheet_by_name("Sheet 1")
#=========================
#list of empty rows
# 1. K-index 1437-1443
# 2. K-index 1437-1443,4855-4856
# 3. K-inndex 760-761,765,769-771,1437-1443
#=========================
def split(word):
    return list(word)
k_ind_val=48
for i in range(2,135):
    cdata=s1.cell(row=i,column=k_ind_val)
    c1 = s1.cell(row=i, column=k_ind_val-8)
    c2 = s1.cell(row=i, column=k_ind_val-7)
    c3 = s1.cell(row=i, column=k_ind_val-6)
    c4 = s1.cell(row=i, column=k_ind_val-5)
    c5 = s1.cell(row=i, column=k_ind_val-4)
    c6 = s1.cell(row=i, column=k_ind_val-3)
    c7 = s1.cell(row=i, column=k_ind_val-2)
    c8 = s1.cell(row=i, column=k_ind_val-1)
    l1=split(cdata.value)
    print(l1," ",i)
    c1.value = l1[0]
    c2.value = l1[2]
    c3.value = l1[4]
    c4.value = l1[6]
    c5.value = l1[8]
    c6.value = l1[10]
    c7.value = l1[12]
    c8.value = l1[14]
wb.save("dataset_2020_new.xlsx")

