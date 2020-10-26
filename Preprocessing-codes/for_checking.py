import openpyxl as op
import numpy as np
'''
wb=op.load_workbook("Final_dataset - nearest vector filling_trial2.xlsx")
#wb=op.load_workbook("Final_dataset - nearest vector filling_trial2.xlsx")
s1=wb.get_sheet_by_name("Sheet1")
for i in range(1,8326):
    for j in range(1,51):
        c=s1.cell(row=i,column=j)
        if c.value=="nk": #or c.value==" " or c.value=="-" or c.value=="- " or c.value=="* " or c.value=="*":
            print(f"({i},{j})")
'''
'''
wb1=op.load_workbook("Final_dataset - nearest vector filling_trial1_modified.xlsx")
wb2=op.load_workbook("Final_dataset - nearest vector filling_trial2_modified.xlsx")
s1=wb1.get_sheet_by_name("Sheet1")
s2=wb2.get_sheet_by_name("Sheet1")
for i in range(1,8382):
    for j in range(1,51):
        c1=s1.cell(row=i,column=j)
        c2=s2.cell(row=i,column=j)
        if c1.value != c2.value:
            print(i,j)
'''

'''
wb1=op.load_workbook("Final_dataset - nearest vector filling_trial1_modified.xlsx")
#wb2=op.load_workbook("Final_dataset - nearest vector filling_trial2_modified.xlsx")
s1=wb1.get_sheet_by_name("Sheet1")
#s2=wb2.get_sheet_by_name("Sheet1")
c_count=0
m_count=0
x_count=0
no_flare_count=0
for i in range(3,8382):
    cc=s1.cell(row=i,column=42)
    cm=s1.cell(row=i,column=43)
    cx=s1.cell(row=i,column=44)
    if cc.value!=0:
        c_count=c_count+1
    if cm.value!=0:
        m_count = m_count + 1
    if cx.value!=0:
        x_count = x_count + 1
print(c_count,m_count,x_count)
'''
'''
wb=op.load_workbook("Final_dataset - nearest vector filling_trial1_modified.xlsx")
s1=wb.get_sheet_by_name("Sheet1")
s_count=0
one_count=0
two_count=0
three_count=0
four_count=0
for i in range(3,8382):
    cs=s1.cell(row=i,column=45)
    c1=s1.cell(row=i,column=46)
    c2=s1.cell(row=i,column=47)
    c3=s1.cell(row=i,column=48)
    c4=s1.cell(row=i,column=49)
    if cs.value!=0:
        s_count=s_count+1
    if c1.value!=0:
        one_count=one_count+1
    if c2.value!=0:
        two_count=two_count+1
    if c3.value!=0:
        three_count=three_count+1
    if c4.value!=0:
        four_count=four_count+1
print(s_count,one_count,two_count,three_count,four_count)
'''
'''
wb=op.load_workbook("Final_dataset - nearest vector filling_trial1_modified.xlsx")
s1=wb.get_sheet_by_name("Sheet1")
max_s=0
max_c=0
max_m=0
max_x=0
max_4=0
for i in range(3,8382):
    cs= s1.cell(row=i,column=45)
    cc = s1.cell(row=i, column=46)
    cm = s1.cell(row=i, column=47)
    cx = s1.cell(row=i, column=48)
    c4 = s1.cell(row=i, column=49)
    if cs.value>max_s:
        max_s=cs.value
    if cc.value>max_c:
        max_c=cc.value
    if cm.value>max_m:
        max_m=cm.value
    if cx.value>max_x:
        max_x=cx.value
    if c4.value > max_4:
        max_4 = c4.value

print(max_s,max_c,max_m,max_x,max_4)
'''