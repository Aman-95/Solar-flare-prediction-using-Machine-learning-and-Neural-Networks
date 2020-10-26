import openpyxl as op

wb=op.load_workbook("Final_dataset - nearest vector filling_trial1_modified.xlsx")
s1=wb.get_sheet_by_name("Sheet1")

# For converting the magnitude of background x-ray flux
# A=1   B=10    C=100

for i in range(3,8382):
    a = s1.cell(row=i,column=7)
    b = s1.cell(row=i,column=8)
    c = s1.cell(row=i,column=9)
    mag=s1.cell(row=i,column=10)
    new=(a.value+(b.value*10)+(c.value)*100)*mag.value
    mag.value=new
wb.save("Dataset_set_1.xlsx")

'''
# For converting the x-ray flares
# C=1    M=100      X=10,000
wb=op.load_workbook("after modifing flare trial 2.xlsx")
s1=wb.get_sheet_by_name("Sheet1")

for i in range(3,8382):
    c = s1.cell(row=i,column=39)
    m = s1.cell(row=i,column=40)
    x = s1.cell(row=i,column=41)
    mag=s1.cell(row=i,column=42)
    mag.value=(c.value+(m.value*100)+(x.value)*10000)
#wb.save("after modifing flare trial 2.xlsx")
'''
'''
# For converting the optical flares
# S=0.1    one=10      two=100      three=1000
for i in range(3,8382):
    s = s1.cell(row=i,column=43)
    one= s1.cell(row=i,column=44)
    two = s1.cell(row=i,column=45)
    three = s1.cell(row=i,column=46)
    four = s1.cell(row=i, column=47)
    mag = s1.cell(row=i,column=48)
    mag.value=((s.value*0.1)+(one.value*10)+(two.value*100)+(three.value*1000)+(four.value*10000))
wb.save("after modifing flare trial 2.xlsx")
'''
