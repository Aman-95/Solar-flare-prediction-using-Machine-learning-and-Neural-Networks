import openpyxl as op
## For fixing radio flux
wb=op.load_workbook("dataset_2020_new.xlsx")
s1=wb.get_sheet_by_name("Sheet 1")
#=========================
#list of empty rows
# 3908-3915,6083,6405,6595
#=========================

def split(word):
    return list(word)
list1=[]
i=0
for i in range(1,134):
    c1 = s1.cell(row=i, column=5)
    c2 = s1.cell(row=i, column=6)
    c3 = s1.cell(row=i, column=7)
    y=c3.value
    list1 = split(y)
    # print(split(y))
    w1 = list1[0]
    list1.pop(0)
    w2="".join(list1)
    #print(w1, " ", i, " ", w2)
    c1.value=w1
    c2.value=w2
wb.save("dataset_2020_new.xlsx")


