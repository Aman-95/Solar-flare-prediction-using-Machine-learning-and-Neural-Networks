import openpyxl as op
wb=op.load_workbook("dataset_2020_new.xlsx")
s1=wb.get_sheet_by_name("Sheet 1")

for i in range(3,8326):
    c=s1.cell(row=i,column=11)
    ca=s1.cell(row=i,column=8)
    cb=s1.cell(row=i,column=9)
    cc=s1.cell(row=i,column=10)
    if c.value=="A":
        ca.value=1
    elif c.value=="B":
        cb.value=1
    else:
        cc.value=1
wb.save("dataset_2020_new.xlsx")